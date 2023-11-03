import tkinter as tk
from tkinter import ttk, filedialog, StringVar, messagebox, simpledialog
import openpyxl
import datetime
import math
import os
import configparser

config = configparser.ConfigParser()
config_file = 'user_preferences.ini'

excel_data = None
edit_round_dialog = None
search_results = []
current_edit_row = None
rounds = [
    {"Round Name": "Round 1", "Start Date": datetime.date(2023, 1, 1), "End Date": datetime.date(2023, 1, 31)},
    {"Round Name": "Round 2", "Start Date": datetime.date(2023, 2, 1), "End Date": datetime.date(2023, 2, 28)},
    {"Round Name": "Round 3", "Start Date": datetime.date(2023, 3, 1), "End Date": datetime.date(2023, 3, 31)},
    {"Round Name": "Round 4", "Start Date": datetime.date(2023, 4, 1), "End Date": datetime.date(2023, 4, 30)}
]

def load_user_preferences():
    global rounds
    if os.path.exists(config_file):
        config.read(config_file)
        for i, round in enumerate(rounds):
            round_name = round["Round Name"]
            if config.has_section(round_name):
                try:
                    start_date = config.get(round_name, 'start_date')
                    end_date = config.get(round_name, 'end_date')
                    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
                    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
                    round["Start Date"] = start_date
                    round["End Date"] = end_date
                except Exception as e:
                    print(f"Error loading preferences for {round_name}: {str(e)}")

def save_user_preferences():
    global rounds
    for round in rounds:
        round_name = round["Round Name"]
        start_date = round["Start Date"].strftime("%Y-%m-%d")
        end_date = round["End Date"].strftime("%Y-%m-%d")
        if not config.has_section(round_name):
            config.add_section(round_name)
        config.set(round_name, 'start_date', start_date)
        config.set(round_name, 'end_date', end_date)

    with open(config_file, 'w') as configfile:
        config.write(configfile)

def read_excel_file():
    global excel_data
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if not file_path:
        return

    try:

        workbook = openpyxl.load_workbook(file_path, read_only=False)

        excel_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            excel_data[sheet_name] = data

        sheet_combobox['values'] = list(excel_data.keys())
        sheet_combobox.current(0)  

        workbook.close()

        search_data()

    except Exception as e:
        tk.messagebox.showerror("Error", f"An error occurred: {str(e)}")

def show_excel_contents():
    global excel_data
    if excel_data:
        selected_sheet = sheet_combobox.get()

        if selected_sheet in excel_data:
            data = excel_data[selected_sheet]

            listbox.delete(0, "end")

            listbox.insert(0, data[0])

            for row in data[1:]:
                listbox.insert("end", row)

def edit_selected_row(event):
    global current_edit_row
    selected_index = listbox.curselection()

    if selected_index:
        current_edit_row = int(selected_index[0])
        entry_fields = []
        current_row_data = listbox.get(current_edit_row)

        edit_dialog = tk.Toplevel(root)
        edit_dialog.title("Edit Row")

        canvas = tk.Canvas(edit_dialog)
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(edit_dialog, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor="nw")

        for i, cell_value in enumerate(current_row_data):
            label = tk.Label(frame, text=f"Column {i + 1}:", font=("Arial", 12))
            label.grid(row=i, column=0, padx=5, pady=5)
            entry_field = tk.Entry(frame, justify="center", font=("Arial", 12))
            entry_field.insert(0, cell_value)
            entry_field.grid(row=i, column=1, padx=5, pady=5)
            entry_fields.append(entry_field)

        button_frame = tk.Frame(frame)
        button_frame.grid(row=len(current_row_data) + 1, columnspan=2, padx=5, pady=10)

        update_button = tk.Button(button_frame, text="Update", command=lambda: update_excel_row(entry_fields, edit_dialog), font=("Arial", 12))
        update_button.grid(row=0, column=0, padx=5)

        cancel_button = tk.Button(button_frame, text="Cancel Admission", command=lambda: cancel_admission(current_edit_row, edit_dialog), font=("Arial", 12))
        cancel_button.grid(row=0, column=1, padx=5)

        # Bind the canvas to the mouse wheel to enable scrolling
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))


def update_round_listbox():
    round_listbox.delete(0, "end")
    for round in rounds:
        round_name = round["Round Name"]
        start_date = round["Start Date"].strftime("%d-%m-%Y")  
        end_date = round["End Date"].strftime("%d-%m-%Y")      
        deduction = ""

        if round_name == "Round 1":
            deduction = "-5000 from installment"
        elif round_name == "Round 2":
            deduction = "-10% from installment"
        elif round_name == "Round 3":
            deduction = "-20% from installment"
        elif round_name == "Round 4":
            deduction = "-50% from installment"

        round_info = f"{round_name} --> Start Date: {start_date} - End Date: {end_date} --> {deduction}"
        round_listbox.insert("end", round_info)

def save_round_dates(dialog):
    global rounds

    for i, round in enumerate(rounds):
        round_name = round["Round Name"]
        start_date_entry = dialog.grid_slaves(row=i, column=1)[0]
        end_date_entry = dialog.grid_slaves(row=i, column=3)[0]

        start_date_str = start_date_entry.get()
        end_date_str = end_date_entry.get()

        try:

            start_date = datetime.datetime.strptime(start_date_str, "%d-%m-%Y").date()
            end_date = datetime.datetime.strptime(end_date_str, "%d-%m-%Y").date()

            if start_date <= end_date:
                round["Start Date"] = start_date
                round["End Date"] = end_date
            else:
                messagebox.showerror("Error", f"End Date for {round_name} must be after Start Date.")
                return
        except ValueError:
            messagebox.showerror("Error", f"Invalid date format for {round_name}. Please use dd-mm-yyyy format.")

    update_round_listbox()
    dialog.destroy()

def cancel_admission(row_index, edit_dialog):
    global excel_data

    if row_index is not None:
        selected_sheet = sheet_combobox.get()

        if selected_sheet in excel_data:

            column_names = excel_data[selected_sheet][0]

            column_select_dialog = tk.Toplevel(root)
            column_select_dialog.title("Select Columns")

            select_label = tk.Label(column_select_dialog, text="Select columns:", font=("Arial", 12))
            select_label.pack(padx=5, pady=5)

            column_listbox = tk.Listbox(column_select_dialog, selectmode=tk.MULTIPLE, font=("Arial", 12))
            column_listbox.pack(padx=5, pady=5)

            for column_name in column_names:
                column_listbox.insert(tk.END, column_name)

            confirm_button = tk.Button(column_select_dialog, text="Confirm", command=lambda: confirm_cancel_admission(row_index, column_listbox.curselection(), edit_dialog, column_select_dialog), font=("Arial", 12))
            confirm_button.pack(padx=5, pady=10)

def confirm_cancel_admission(row_index, selected_columns, edit_dialog, column_select_dialog):
    global excel_data, rounds

    if row_index is not None:
        selected_sheet = sheet_combobox.get()

        if selected_sheet in excel_data:

            row_data = list(excel_data[selected_sheet][row_index])

            deduction = {}

            try:

                selected_column_data = []

                total_amount = sum(row_data[col_index] for col_index in selected_columns if row_data[col_index])

                selected_round_name = simpledialog.askstring("Select Round", "Select a round:", initialvalue="Round 1")

                if selected_round_name:
                    selected_round = next((round for round in rounds if round["Round Name"] == selected_round_name), None)

                    if selected_round:

                        for col_index in selected_columns:
                            cell_value = row_data[col_index]

                            if isinstance(cell_value, str) and not cell_value.replace('.', '', 1).isdigit():

                                cell_value = float(cell_value.replace(',', '').replace(' ', ''))
                            else:
                                cell_value = float(cell_value)

                            if cell_value is not None and not math.isnan(cell_value):
                                column_name = excel_data[selected_sheet][0][col_index]
                                deduction[column_name] = 0

                                if selected_round_name == "Round 1":
                                    deduction[column_name] = 5000
                                elif selected_round_name == "Round 2":
                                    deduction[column_name] = cell_value * 0.10
                                elif selected_round_name == "Round 3":
                                    deduction[column_name] = cell_value * 0.20
                                elif selected_round_name == "Round 4":
                                    deduction[column_name] = cell_value * 0.50

                                selected_column_data.append(f"{column_name}: {cell_value} - Deduction: {deduction[column_name]}")

                        excel_data[selected_sheet][row_index] = tuple(row_data)

                        refund_amount = total_amount - sum(deduction.values())

                        selected_round = next((round for round in rounds if round["Round Name"] == selected_round_name), None)
                        if selected_round:
                            round_name = selected_round_name
                            start_date = selected_round["Start Date"]
                            end_date = selected_round["End Date"]
                        else:
                            round_name = "N/A"
                            start_date = "N/A"
                            end_date = "N/A"

                        message = f"Total Amount: {total_amount}\nRefund Amount for {selected_round_name}: {refund_amount}\n\nSelected Column Data:\n"
                        message += "\n".join(selected_column_data)
                        messagebox.showinfo("Calculation Result", message)

                        save_calculation_result(selected_sheet, row_data, total_amount, deduction, selected_column_data, message, refund_amount, round_name, start_date, end_date)

                    else:
                        messagebox.showerror("Error", "Invalid round selection")
                else:
                    messagebox.showerror("Error", "Round selection canceled")
            except ValueError:
                messagebox.showerror("Error", "Invalid column selection")
            finally:
                edit_dialog.destroy()  
                column_select_dialog.destroy()  

                show_excel_contents()

def save_calculation_result(sheet_name, row_data, total_amount, deduction, selected_column_data, message, refund_amount, round_name, start_date, end_date):
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
    if not file_path:
        return
    
    deduction_info = {
        "Round 1": "-5000 from installment",
        "Round 2": "-10% from installment",
        "Round 3": "-20% from installment",
        "Round 4": "-50% from installment",
    }

    with open(file_path, "a") as file:
        file.write("Calculation Result\n\n")
        file.write(f"Sheet Name: {sheet_name}\n")
        file.write(f"Row Data: {row_data}\n")
        file.write(f"Total Amount: {total_amount}\n")
        file.write("\nSelected Column Data:\n")
        for data in selected_column_data:
            column_name, column_value, column_deduction = data.split(": ")
            file.write(f"- {column_name}: {column_value} : Deducted: {column_deduction}\n")
        file.write(f"Refund Amount: {refund_amount}\n")
        file.write(f"Round Name: {round_name} ({deduction_info.get(round_name, 'N/A')})\n")
        file.write(f"Round Start Date: {start_date.strftime('%d-%m-%Y')}\n")
        file.write(f"Round End Date: {end_date.strftime('%d-%m-%Y')}\n")

    messagebox.showinfo("File Saved", f"Calculation result saved as {file_path}")


def open_edit_round_dialog():
    global edit_round_dialog

    edit_round_dialog = tk.Toplevel(root)
    edit_round_dialog.title("Update Round Dates")

    for i, round in enumerate(rounds):
        round_name = round["Round Name"]
        start_date = round["Start Date"]
        end_date = round["End Date"]

        round_label = tk.Label(edit_round_dialog, text=f"{round_name} ({start_date} - {end_date}):")
        round_label.grid(row=i, column=0, padx=5, pady=5)

        start_date_entry = tk.Entry(edit_round_dialog, justify="center")
        start_date_entry.insert(0, start_date.strftime("%d-%m-%Y"))
        start_date_entry.grid(row=i, column=1, padx=5, pady=5)

        end_date_entry = tk.Entry(edit_round_dialog, justify="center")
        end_date_entry.insert(0, end_date.strftime("%d-%m-%Y"))
        end_date_entry.grid(row=i, column=3, padx=5, pady=5)

    save_button = tk.Button(edit_round_dialog, text="Save", command=save_round_dates)
    save_button.grid(row=len(rounds), columnspan=2, padx=5, pady=10)

def save_round_dates():
    global rounds, edit_round_dialog

    for i, round in enumerate(rounds):
        round_name = round["Round Name"]
        start_date_entry = edit_round_dialog.grid_slaves(row=i, column=1)[0]
        end_date_entry = edit_round_dialog.grid_slaves(row=i, column=3)[0]

        start_date_str = start_date_entry.get()
        end_date_str = end_date_entry.get()

        try:

            start_date = datetime.datetime.strptime(start_date_str, "%d-%m-%Y").date()
            end_date = datetime.datetime.strptime(end_date_str, "%d-%m-%Y").date()

            if start_date <= end_date:
                round["Start Date"] = start_date
                round["End Date"] = end_date
            else:
                messagebox.showerror("Error", f"End Date for {round_name} must be after Start Date.")
                return
        except ValueError:
            messagebox.showerror("Error", f"Invalid date format for {round_name}. Please use dd-mm-yyyy format.")

    update_round_listbox()
    edit_round_dialog.destroy()

def update_excel_row(entry_fields, edit_dialog):
    global current_edit_row, excel_data

    if current_edit_row is not None:
        selected_sheet = sheet_combobox.get()
        edited_row_values = []

        for entry_field in entry_fields:
            edited_row_values.append(entry_field.get())

        if selected_sheet in excel_data:
            excel_data[selected_sheet][current_edit_row] = tuple(edited_row_values)

            listbox.delete(0, "end")
            data = excel_data[selected_sheet]
            listbox.insert(0, data[0])

            for row in data[1:]:
                listbox.insert("end", row)

            current_edit_row = None
            edit_dialog.destroy()  

def search_data(event=None):
    global excel_data, search_results
    if excel_data:
        search_term = search_var.get().strip().lower()

        listbox.delete(0, "end")

        search_results = []
        for sheet_name, data in excel_data.items():
            sheet_data = list(data[0])  
            search_results.append([f"Sheet: {sheet_name}"] + sheet_data)  
            for row in data[1:]:

                filtered_row = [cell if cell is not None and cell != "" else "-" for cell in row]
                if any(search_term in str(cell).strip().lower() for cell in row):
                    search_results.append([f"Sheet: {sheet_name}"] + filtered_row)

        for row in search_results:
            listbox.insert("end", row)

root = tk.Tk()
root.title("Fees Deduction System")
root.geometry("1000x800")

round_info_frame = tk.LabelFrame(root, text="Round Information", font=("Verdana", 14))
round_info_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

round_listbox = tk.Listbox(round_info_frame, font=("Verdana", 16))
round_listbox.pack(fill='both', expand=True, padx=10, pady=5)

update_round_button = tk.Button(round_info_frame, text="Update Round Dates", command=open_edit_round_dialog, font=("Verdana", 12))
update_round_button.pack(side="top", padx=10, pady=5, anchor='w')

excel_frame = tk.LabelFrame(root, text="Excel File Operations", font=("Verdana", 14))
excel_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

import_button = tk.Button(excel_frame, text="Import Excel File", command=read_excel_file, font=("Verdana", 12))
import_button.grid(row=0, column=0, padx=10, pady=5)

show_button = tk.Button(excel_frame, text="Show Excel Contents", command=show_excel_contents, font=("Verdana", 12))
show_button.grid(row=0, column=1, padx=10, pady=5)

search_frame = tk.LabelFrame(root, text="Search and Data Display", font=("Verdana", 14))
search_frame.grid(row=2, column=0, padx=10, pady=5, sticky="ew")

search_var = StringVar()  
search_var.trace_add("write", lambda *args: search_data())  

search_label = tk.Label(search_frame, text="Search:", font=("Verdana", 12))
search_label.grid(row=0, column=0)

search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Verdana", 12))
search_entry.grid(row=0, column=1, sticky="ew")

listbox_frame = tk.LabelFrame(root, text="Sheet Data and Search Results", font=("Verdana", 14))
listbox_frame.grid(row=3, column=0, padx=10, pady=5, sticky="nsew")

listbox = tk.Listbox(listbox_frame, font=("Verdana", 12))
listbox.pack(fill="both", expand=True, padx=10, pady=5)

listbox.bind("<Double-Button-1>", edit_selected_row)

sheet_combobox = ttk.Combobox(listbox_frame, values=[], state="readonly", font=("Verdana", 12))
sheet_combobox.pack(side="top", padx=10, pady=5)

root.columnconfigure(0, weight=1)
root.rowconfigure(3, weight=1)
listbox_frame.columnconfigure(0, weight=1)

def close_application():
    save_user_preferences()  
    root.destroy()

root.protocol("WM_DELETE_WINDOW", close_application)

load_user_preferences()

root.mainloop()